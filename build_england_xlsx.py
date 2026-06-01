#!/usr/bin/env python3
"""England-only promotion/relegation table (tiers 1-5), splitting direct vs
playoff, with direct source links per row."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["Division", "Tier", "Clubs",
           "Promoted: direct", "Promoted: playoff", "Promotion %",
           "Relegated: direct", "Relegated: playoff", "Relegation %",
           "Playoff format", "Source 1", "Source 2"]

# division, tier, clubs, p_dir, p_po, p%, r_dir, r_po, r%, playoff_format,
#   (s1_text, s1_url), (s2_text, s2_url)
ROWS = [
    ["Premier League", 1, "20", "—", "—", "—", "3", "0", "15.0%",
     "No play-off; bottom 3 relegated automatically",
     ("premierleague.com - explained", "https://www.premierleague.com/en/premier-league-explained"),
     ("Wikipedia - English football league system", "https://en.wikipedia.org/wiki/English_football_league_system")],
    ["EFL Championship", 2, "24", "2", "1", "12.5%", "3", "0", "12.5%",
     "Top 2 up automatically; 3rd-6th play off (two-legged semis + Wembley final) for 1 place",
     ("Wikipedia - EFL Championship", "https://en.wikipedia.org/wiki/EFL_Championship"),
     ("EFL - Sky Bet Play-Offs", "https://www.efl.com/competitions/sky-bet-play-offs/")],
    ["EFL League One", 3, "24", "2", "1", "12.5%", "4", "0", "16.7%",
     "Top 2 up automatically; 3rd-6th play off for 1 place; bottom 4 relegated",
     ("Wikipedia - EFL League One", "https://en.wikipedia.org/wiki/EFL_League_One"),
     ("EFL - Sky Bet Play-Offs", "https://www.efl.com/competitions/sky-bet-play-offs/")],
    ["EFL League Two", 4, "24", "3", "1", "16.7%", "2", "0", "8.3%",
     "Top 3 up automatically; 4th-7th play off for 1 place; bottom 2 relegated",
     ("Wikipedia - EFL League Two", "https://en.wikipedia.org/wiki/EFL_League_Two"),
     ("Wikipedia - EFL League Two play-offs", "https://en.wikipedia.org/wiki/EFL_League_Two_play-offs")],
    ["National League", 5, "24", "1", "1", "8.3%", "4", "0", "16.7%",
     "Champion up automatically; 2nd-7th play off (eliminators + semis + Wembley final) for 1 place; bottom 4 relegated to NL North/South. Promotion conditional on EFL ground-grading.",
     ("Wikipedia - National League (division)", "https://en.wikipedia.org/wiki/National_League_(division)"),
     ("Wikipedia - National League (English football)", "https://en.wikipedia.org/wiki/National_League_(English_football)")],
]

FOOTNOTES = [
    "Scope: England tiers 1-5, 2024-25 / 2025-26 structure. 'Direct' = automatic on final standings; 'playoff' = via end-of-season play-offs. % = (direct + playoff) / clubs x 100.",
    "England has NO relegation play-offs anywhere in tiers 1-5 - relegation is purely on final position.",
    "Coming change: from 2026-27 the Championship play-offs expand from 3rd-6th (4 clubs) to 3rd-8th (6 clubs); still 1 promoted (source: Wikipedia EFL Championship).",
    "National League relegation is normally 4 but the FA can adjust it if an EFL club is expelled; promotion to the EFL also requires meeting EFL ground-grading (Category A, min 4,000 capacity).",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
PROMO_HDR = PatternFill("solid", fgColor="2E7D32")
RELEG_HDR = PatternFill("solid", fgColor="C0392B")
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
LINK_FONT = Font(color="0563C1", underline="single")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ROW_FILL = PatternFill("solid", fgColor="FCE4D6")

wb = Workbook()
ws = wb.active
ws.title = "England P&R"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
t = ws.cell(row=1, column=1,
            value="England - Promotion & Relegation (direct vs play-off), Tiers 1-5 (2024-25 / 2025-26)")
t.font = TITLE_FONT
ws.row_dimensions[1].height = 20

for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.alignment = CENTER
    cell.border = BORDER
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PROMO_HDR if c in (4, 5, 6) else RELEG_HDR if c in (7, 8, 9) else HEADER_FILL
ws.row_dimensions[2].height = 30

r = 3
for row in ROWS:
    for c in range(1, len(HEADERS) + 1):
        if c in (11, 12):  # source columns
            text, url = row[c - 1]
            cell = ws.cell(row=r, column=c, value=text)
            cell.hyperlink = url
            cell.font = LINK_FONT
            cell.alignment = LEFT
        else:
            cell = ws.cell(row=r, column=c, value=row[c - 1])
            cell.alignment = LEFT if c in (1, 10) else CENTER
            if c == 1:
                cell.font = Font(bold=True)
        cell.fill = ROW_FILL
        cell.border = BORDER
    ws.row_dimensions[r].height = 46
    r += 1

widths = [20, 5, 8, 13, 14, 11, 13, 14, 11, 40, 30, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"
ws.sheet_view.showGridLines = False

r += 1
for note in FOOTNOTES:
    ws.cell(row=r, column=1, value="Note: " + note).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
    r += 1

OUT = "england_promotion_relegation.xlsx"
wb.save(OUT)
print("Wrote", OUT)
