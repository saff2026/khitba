#!/usr/bin/env python3
"""Generate an Excel workbook comparing promotion/relegation across the
first 5 tiers of the football pyramids in England, Spain, Germany, Italy
and France (2024-25 / 2025-26 league structures)."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Data: one row per division. Fields match the user's request.
#   country, division, tier, clubs, promoted, promo_pct, relegated, rel_pct,
#   notes, sources
# ---------------------------------------------------------------------------

HEADERS = [
    "Country", "Division", "Tier", "Clubs in division",
    "Promoted / season (auto + playoff)", "Promotion %",
    "Relegated / season (auto + playoff)", "Relegation %",
    "Notes (playoffs, conditional P/R, reserve-team & regional-group rules)",
    "Sources",
]

ROWS = [
    # ---------------- ENGLAND ----------------
    ["England", "Premier League", 1, "20", "— (top tier)", "—",
     "3 (3 auto + 0 PO)", "15.0%",
     "No relegation play-off; bottom 3 down. No reserve/B-teams anywhere in the pyramid.",
     "Wikipedia 2024-25 Premier League"],
    ["England", "EFL Championship", 2, "24", "3 (2 auto + 1 PO)", "12.5%",
     "3 (3 auto + 0 PO)", "12.5%",
     "PO contested by 3rd-6th, 1 promoted (Wembley final). From 2026-27 the PO expands to 3rd-8th (6 clubs), still 1 promoted. No reserve teams in EFL.",
     "Wikipedia 2024-25 EFL Championship; EFL Championship play-offs (Wikipedia)"],
    ["England", "EFL League One", 3, "24", "3 (2 auto + 1 PO)", "12.5%",
     "4 (4 auto + 0 PO)", "16.7%",
     "PO contested by 3rd-6th, 1 promoted.",
     "Wikipedia 2024-25 EFL League One"],
    ["England", "EFL League Two", 4, "24", "4 (3 auto + 1 PO)", "16.7%",
     "2 (2 auto + 0 PO)", "8.3%",
     "PO contested by 4th-7th, 1 promoted. Lowest fully-professional tier.",
     "Wikipedia 2024-25 EFL League Two"],
    ["England", "National League (Premier)", 5, "24", "2 (1 auto + 1 PO)", "8.3%",
     "~4 (4 auto + 0 PO)", "16.7%",
     "PO contested by 2nd-7th (6 clubs), 1 promoted. Promotion to EFL is CONDITIONAL on ground-grading (EFL Cat. A, min 4,000 capacity); a champion failing it can be denied. Tier 6 below (NL North/South) is regionalized; tier 5 itself is national. Relegation normally 4 but FA can adjust if an EFL club is expelled. No reserve/B-teams permitted.",
     "Wikipedia 2024-25 National League; National League (English football) (Wikipedia); FA National League System regs & ground-grading handbook"],

    # ---------------- SPAIN ----------------
    ["Spain", "LaLiga (EA Sports)", 1, "20", "— (top tier)", "—",
     "3 (3 auto + 0 PO)", "15.0%",
     "No relegation play-off. Reserve teams (filiales) can NEVER reach this tier.",
     "Wikipedia 2024-25 & 2025-26 La Liga"],
    ["Spain", "LaLiga2 (Hypermotion)", 2, "22", "3 (2 auto + 1 PO)", "13.6%",
     "4 (4 auto + 0 PO)", "18.2%",
     "PO contested by 3rd-6th, 1 promoted. CEILING for all B-teams: a B-team in the top 6 cannot be promoted; the slot reroutes to the next eligible club.",
     "Wikipedia 2024-25 & 2025-26 Segunda Division"],
    ["Spain", "Primera Federacion", 3, "40 (2 groups x 20)", "4 (2 auto + 2 PO)", "10.0%",
     "10 (10 auto + 0 PO)", "25.0%",
     "Per group: 1 champion auto + 1 PO winner (from 2nd-5th). Per-group promotion 10%, relegation 25% (bottom 5 of each group). RFEF-run; B-team rule applies (ineligible berths reroute).",
     "Wikipedia 2024-25 Primera Federacion; 2024 Primera Federacion play-offs"],
    ["Spain", "Segunda Federacion", 4, "90 (5 groups x 18)", "10 (5 auto + 5 PO)", "11.1%",
     "~27 (25 auto + ~2 PO)", "~30%",
     "Per group: 1 champion auto + 1 PO winner (2nd-5th). Relegation = 5 auto/group + ~2 via a NATIONAL mini-playoff among the weakest 13th-placed sides, so totals vary. B-team rule applies.",
     "Wikipedia 2024-25 & 2025-26 Segunda Federacion; 2024 Segunda Federacion play-offs"],
    ["Spain", "Tercera Federacion", 5, "~325 (18 regional groups)", "27 (18 auto + 9 PO)", "~8.3%",
     "~54+ (regionally set)", "~17%+",
     "18 group champions auto + 9 via play-offs (2nd-5th). The 9 PO places do not divide evenly per group (per-group figures are averages). Relegation count is set INDEPENDENTLY by each autonomous-community federation (~3-5/group) and varies by group and year. B-team rule applies.",
     "Wikipedia 2024-25 & 2025-26 Tercera Federacion; Spanish football league system (Wikipedia)"],

    # ---------------- GERMANY ----------------
    ["Germany", "Bundesliga", 1, "18", "— (top tier)", "—",
     "3 (2 auto + 1 PO)", "11.1-16.7%",
     "Relegation play-off (Relegation): 16th vs 2. Bundesliga 3rd, two legs, no away goals since 2021-22. Auto alone = 11.1%; with PO loss = 16.7%.",
     "Bundesliga.com (official) P/R FAQ; DFL.de 2025 relegation play-off schedule"],
    ["Germany", "2. Bundesliga", 2, "18", "3 (2 auto + 1 PO)", "11.1-16.7%",
     "3 (2 auto + 1 PO)", "11.1-16.7%",
     "Two play-offs touch this tier: up (3rd vs BL 16th) and down (16th vs 3. Liga 3rd). Highest tier a reserve ('II') team CANNOT reach.",
     "Wikipedia 2. Bundesliga; Bundesliga.com"],
    ["Germany", "3. Liga", 3, "20", "3 (2 auto + 1 PO)", "10.0-15.0%",
     "4 (4 auto + 0 PO)", "20.0%",
     "PO: 3rd vs 2. Bundesliga 16th. Reserve ('II') teams may PLAY here but CANNOT be promoted (slot passes to next eligible club); reserves also barred from DFB-Pokal. Single national league (DFB-run).",
     "Wikipedia 2024-25 3. Liga; Reserve team (Wikipedia)"],
    ["Germany", "Regionalliga", 4, "~90 (5 groups x 18)", "~4 of 5 champions (~3 direct + 1 PO)", "~4.4% (national)",
     "varies (~2-4 / group)", "~10-20% / group",
     "ANOMALY: 5 regional champions (Nord, Nordost, West, Sudwest, Bayern) but only ~4 promotion slots, so ONE champion misses out each year. Sudwest & West usually direct; 3rd direct slot rotates; the other two contest a play-off. 2024-25: Sudwest/West/Bayern direct, Nord-Nordost PO. 4-division reform approved in principle (earliest 2028-29). Reserve teams cannot be promoted.",
     "Wikipedia 2024-25 & 2025-26 Regionalliga; Promotion to the 3. Liga (Wikipedia); World Football Index"],
    ["Germany", "Oberliga", 5, "varies (14 parallel leagues, ~16-18 each)", "~1 champion / league (some via PO)", "~6% / league",
     "varies (regional)", "varies",
     "Run by state associations, so club counts and P/R numbers are set annually by each region's promotion/relegation key. Same reserve-team promotion bar applies.",
     "Wikipedia 2024-25 Oberliga; Oberliga (football) (Wikipedia)"],

    # ---------------- ITALY ----------------
    ["Italy", "Serie A", 1, "20", "— (top tier)", "—",
     "3 (3 auto + 0 PO)", "15.0%",
     "Cleanest tier; no playout. Reserve teams barred from Serie A.",
     "Wikipedia 2025-26 Serie A; Serie A (Wikipedia)"],
    ["Italy", "Serie B", 2, "20", "3 (2 auto + 1 PO)", "15.0%",
     "4 (3 auto + 1 playout)", "20.0%",
     "PO contested by 3rd-8th, 1 promoted. PLAYOUT (16th v 17th) decides the 4th relegation, but is CANCELLED if 16th finishes >=4 pts clear (17th relegated automatically).",
     "Wikipedia 2024-25 Serie B; Football-Italia (playout); Destination Calcio"],
    ["Italy", "Serie C", 3, "60 (3 groups x 20)", "4 (3 auto + 1 large PO)", "6.7%",
     "~9 (3 auto + ~6 playout)", "~15%",
     "3 group champions auto + ONE promotion from a ~28-team national play-off bracket (lowest yield-per-participant anywhere). Relegation = 1 auto/group + playouts. Reserve/U23 teams allowed (Juventus Next Gen 2018, Atalanta U23 2023). Heavy ripescaggio (see workbook 'Cross-cutting notes').",
     "Wikipedia 2024-25 Serie C; nss-sports; Cult of Calcio; Juventus Next Gen (Wikipedia)"],
    ["Italy", "Serie D", 4, "~162 (9 groups x ~18)", "9 (9 auto + 0 PO)", "~5.6%",
     "~36 (~9 auto + ~27 playout)", "~22%",
     "9 group champions promoted automatically (runner-up play-offs only set ripescaggio priority). Top amateur tier (LND). Mid-season withdrawals/exclusions reshuffle relegation math.",
     "Wikipedia 2024-25 & 2025-26 Serie D"],
    ["Italy", "Eccellenza", 5, "~474 (~28 regional groups)", "~36 (~28 champions + ~7 PO + 1 cup)", "~7.6%",
     "region-set (~2-4 / group)", "~12-20%",
     "Fully regionalized: each regional committee sets its own group sizes, play-off bracket and relegation count. Coppa Italia Dilettanti winner also promoted. Reserve teams permitted.",
     "Wikipedia Eccellenza; Italian football league system"],

    # ---------------- FRANCE ----------------
    ["France", "Ligue 1 (McDonald's)", 1, "18", "— (top tier)", "—",
     "2 auto + 1 barrage (16th)", "11.1-16.7%",
     "18 clubs since 2023-24. Barrage: 16th vs Ligue 2 PO winner (2 legs). DNCG (financial regulator) can administratively relegate clubs regardless of position (e.g. Lyon provisionally relegated 2025, reinstated on appeal).",
     "Wikipedia 2024-25 & 2025-26 Ligue 1; Ligue 1 relegation/promotion play-offs; ESPN (Lyon); DNCG (Wikipedia)"],
    ["France", "Ligue 2", 2, "18", "2 auto + 1 barrage", "11.1-16.7%",
     "2 auto + 1 barrage", "11.1-16.7%",
     "PO winner (~3rd-5th) plays Ligue 1 16th; relegation barrage 16th vs National 3rd (reintroduced 2024-25 with the 20->18 downsizing). DNCG can force relegation (Bordeaux dropped two tiers to National 2 in 2024-25).",
     "Wikipedia Ligue 2; 2024-25 & 2025-26 Ligue 2; France 3 (Bordeaux DNCG)"],
    ["France", "Championnat National", 3, "18", "2 auto + 1 barrage", "11.1-16.7%",
     "2 (auto, transitional)", "~11.1%",
     "3rd plays Ligue 2 16th in the barrage. Relegation reduced to ~2 during downsizing (historically 3-4). Becomes professional 'Ligue 3' from 2026-27.",
     "Wikipedia Championnat National; 2024-25 Championnat National; FFF official Ligue 3 launch"],
    ["France", "Championnat National 2", 4, "48 (3 groups x 16)", "up to 3 (group champions)*", "~6.25%",
     "~10 in 2024-25 (restructuring)", "~20.8%",
     "*3 regional groups (not 4). Reserve teams may WIN a group but CANNOT be promoted to National (tier 3); the slot reroutes to the next eligible club. Geographic groups -> relegation counts vary by group. Renamed 'National 1' from 2026-27.",
     "Wikipedia Championnat National 2; 2024-25 Championnat National 2; FFF N1/N2 regs"],
    ["France", "Championnat National 3", 5, "140 -> 112 (10 -> 8 groups x 14)", "1 per group (10 -> 8)*", "~7.1% / group",
     "many (variable, regional)", "variable / elevated",
     "*Run by the regional leagues; mid-downsizing (140->112) temporarily inflates relegation. Group winner promoted (next eligible if ineligible). Reserve teams capped at tier 4. Renamed 'National 2' from 2026-27.",
     "Wikipedia Championnat National 3; 2024-25 Championnat National 3; French football league system"],
]

CHURN_HEADERS = ["Metric", "Highest churn", "Lowest churn", "Comment"]
CHURN_ROWS = [
    ["Promotion % (share moving up)",
     "England League Two 16.7%; Italy Serie B 15%; Eng Championship/League One 12.5%",
     "Germany Regionalliga ~4.4% (national); Italy Serie D ~5.6% / Serie C 6.7%; Eng National League 8.3%",
     "Upper-mid professional tiers move the most clubs up relative to size; vast regionalized tiers feeding one national league above move the fewest."],
    ["Relegation % (share dropped)",
     "Spain Segunda Federacion ~30%; Spain Primera Federacion 25%; Italy Serie D ~22%; Germany 3. Liga 20%; Italy Serie B 20%",
     "England League Two 8.3% (only 2 down); France National/Ligue tiers ~11% automatic floor",
     "Spain's RFEF tiers are the harshest (a quarter to a third relegated each year) due to funneling many regional groups into fewer national divisions."],
    ["Most volatile pyramid", "Spain (15% -> 18% -> 25% -> 30% relegation gradient)", "England (stable 24-club tiers)",
     "Spain has by far the steepest relegation gradient down the tiers."],
    ["Most play-off dependent (top)", "Germany (every tier 1/2/3 boundary = 2 auto + 1 relegation play-off)", "England (no relegation play-offs at all)",
     "Germany also has the most contentious structural step at tier 4 (5 champions, ~4 slots)."],
    ["Least predictable", "Italy (ripescaggio / financial exclusions routinely override nominal numbers)", "England (transparent, rule-clean)",
     "Italian actual P/R frequently differs from the rules due to financial exclusions and 'fishing back'."],
    ["Most administratively interventionist", "France (DNCG can relegate clubs for finances)", "England (sporting results only)",
     "France is also mid-restructuring (18-club top flights; Ligue 3 arrives 2026-27)."],
    ["Most stable single tier", "Premier League / Championship symmetry (15% & 12.5% both ways)", "", "England."],
    ["Most extreme single tier", "Spain Segunda Federacion (~30% relegated / year)", "", "Spain."],
]

NOTES = [
    "Scope: most recent available league structures (2024-25 / 2025-26 seasons).",
    "'auto' = automatic on final league standings; 'PO' = via play-off; 'barrage' = French two-legged promotion/relegation play-off; 'playout' = Italian two-legged relegation play-off.",
    "Promotion % = clubs promoted / clubs in division x 100. Relegation % = clubs relegated / clubs in division x 100. For regionalized tiers, percentages are computed per group where appropriate; national totals are given in the cell.",
    "Ranges are given where the count depends on play-offs/barrages, licensing, reserve-team eligibility, financial exclusions, or ongoing league restructuring.",
    "Reserve/B-team rules: Spain - filiales capped at LaLiga2 and auto-relegated if parent drops into their tier; Germany - 'II' teams play up to 3. Liga but cannot be promoted above it (and are barred from the DFB-Pokal); Italy - reserve/U23 teams allowed only from Serie C down (never Serie A); France - reserves capped at National 2 (cannot reach tier 3+); England - no reserve teams in the senior pyramid.",
    "Conditional P/R: England - EFL ground-grading can deny a National League champion promotion; Italy - ripescaggio (financial exclusions refilled by readmitting/elevating clubs) routinely changes actual numbers; France - the DNCG can administratively relegate clubs for financial reasons (e.g. Bordeaux 2024-25, Lyon 2025 overturned on appeal).",
    "Regionalized tiers: Spain tiers 3-5, Germany tiers 4-5, Italy tiers 3-5, France tiers 4-5. Club counts and lower-tier relegation totals are set annually by regional federations and are therefore approximate.",
    "Upcoming changes flagged but outside the requested seasons: England Championship play-offs expand to 3rd-8th from 2026-27; France launches professional 'Ligue 3' (and renames National 2 -> National 1, National 3 -> National 2) from 2026-27; Germany's Regionalliga 4-division reform approved in principle, earliest 2028-29.",
    "Sourcing note: primarily official league/federation sites (Premier League/EFL, LaLiga/RFEF, Bundesliga.com/DFL/DFB, Lega Serie A/B/Pro/LND, LFP/FFF) and Wikipedia season/structure articles, cross-checked against reputable football media. Several official sites blocked direct fetching, so those figures were verified across multiple independent sources.",
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COUNTRY_FILLS = {
    "England": "FCE4D6", "Spain":  "FFF2CC", "Germany": "E2EFDA",
    "Italy":   "DEEBF7", "France": "EDE7F6",
}

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# --- Sheet 1: full comparison ---
ws = wb.active
ws.title = "P&R Comparison"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
t = ws.cell(row=1, column=1,
            value="Promotion & Relegation - Top 5 European Football Pyramids, Tiers 1-5 (2024-25 / 2025-26)")
t.font = TITLE_FONT
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, h in enumerate(HEADERS, 1):
    ws.cell(row=2, column=c, value=h)
style_header(ws, 2, len(HEADERS))

r = 3
for row in ROWS:
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = CENTER if c in (3, 6, 8) else WRAP_TOP
        cell.fill = PatternFill("solid", fgColor=COUNTRY_FILLS[row[0]])
        if c == 1:
            cell.font = Font(bold=True)
    r += 1

widths = [11, 24, 6, 18, 22, 11, 22, 12, 60, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"
ws.sheet_view.showGridLines = False

# --- Sheet 2: churn summary ---
ws2 = wb.create_sheet("Churn Summary")
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CHURN_HEADERS))
t2 = ws2.cell(row=1, column=1, value="Where promotion/relegation churn is highest and lowest")
t2.font = TITLE_FONT
for c, h in enumerate(CHURN_HEADERS, 1):
    ws2.cell(row=2, column=c, value=h)
style_header(ws2, 2, len(CHURN_HEADERS))
r = 3
for row in CHURN_ROWS:
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = WRAP_TOP
        if c == 1:
            cell.font = Font(bold=True)
    r += 1
for i, w in enumerate([34, 42, 42, 50], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.sheet_view.showGridLines = False

# --- Sheet 3: methodology & notes ---
ws3 = wb.create_sheet("Notes & Methodology")
t3 = ws3.cell(row=1, column=1, value="Methodology, definitions & cross-cutting notes")
t3.font = TITLE_FONT
ws3.column_dimensions["A"].width = 120
for i, n in enumerate(NOTES, start=3):
    cell = ws3.cell(row=i, column=1, value="• " + n)
    cell.alignment = WRAP_TOP
ws3.sheet_view.showGridLines = False

OUT = "football_promotion_relegation_top5.xlsx"
wb.save(OUT)
print("Wrote", OUT)
